from openpyxl import load_workbook
import openpyxl

load_wb = load_workbook("C:/Users/AORUS/Desktop/Project/crawling/crawling.xlsx", data_only=True)
load_ws = load_wb['Sheet']

print(load_ws['A1'].value)
print(load_ws.cell(1,2).value)

wb = openpyxl.load_workbook('crawling.xlsx')
sheet1 = wb.active
x=0
sheet1.title = 'Sheet'
for num in range(0, 10):
    x += 1
    sheet1.cell(x,3, num)

wb.save('crawling1.xlsx')